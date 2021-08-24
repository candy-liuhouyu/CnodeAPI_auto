"""
@Author : candy_liu
@Contact :candy_liuhouyu@163.com
@Date : 2021/08/19 18:34
@Desc :封装读取数据的方法
"""

import json
from openpyxl import load_workbook
import action


# 读取TXT文件的数据
def read_txt(filename):
    filepath = action.BASE_DIR + "/data/" + filename
    with open(filepath, "r", encoding="utf-8") as f:
        return f.readlines()


def load_json(file_path):
    # json数据驱动
    data = json.load(open(file_path, mode='r', encoding='utf8'))
    json_data = data['test_data']
    return json_data


# 读取excel文件的数据
def read_excel(file_path, sheet_name):
    # 使用openpyxl读取Excel的数据demo
    wb = load_workbook(file_path)  # 加载文件目录
    # print(wb.worksheets)  # 获取Excel的sheet表
    ws = wb[sheet_name]  # 只需要取其中一个sheet表的数据
    # print(ws['B1'].value)   # 打印sheet表中的b1数据
    # print(len(tuple(ws.rows)))
    test_data = []
    for x in range(2, len(tuple(ws.rows)) + 1):
        testcase_data = []
        for y in range(2, len(tuple(ws.columns)) + 1):
            testcase_data.append(ws.cell(row=x, column=y).value)  # 取出单元格的数据存放到列表中
            # print(ws.cell(row=x, column=y).value)
        test_data.append(testcase_data)  # 将每行的值存放到列表中心
    return test_data


if __name__ == '__main__':
    test_data = load_json("D:/document/MyProject/data/topics.json")
    print(test_data)